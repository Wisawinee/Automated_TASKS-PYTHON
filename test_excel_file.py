from openpyxl import load_workbook

excelfile = load_workbook('sample.xlsx')
sheet = excelfile['DataBag']
maxrow = sheet.max_row
maxrow = maxrow+1

for i in range(2,maxrow):
    txt_product_name = sheet.cell(row=i,column=1).value
    txt_product_quantity = sheet.cell(row=i,column=2).value
    txt_product_price = sheet.cell(row=i,column=3).value
    print(txt_product_name,txt_product_quantity,txt_product_price)


'''from openpyxl import Workbook
wb = Workbook()
# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = 42
# Rows can also be appended
ws.append([1, 2, 3])
# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()
# Save the file
wb.save("sample.xlsx")'''