from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook

driver = webdriver.Chrome("chromedriver.exe")
def login_to_web(url='https://noraphat.dev/POS_NAME/index.php'):
    driver.get(url)

    txt_username = driver.find_element_by_name('username')
    txt_username.send_keys('admin')

    txt_password = driver.find_element_by_name('password')
    txt_password.send_keys('1234')

    txt_bntLogin = driver.find_element_by_name('bntLogin')
    txt_bntLogin.click()

def add_product(url,exc_product_name,exc_product_quantity,exc_product_price):
    driver.get(url)

    txt_product_name = driver.find_element_by_name('product_name')
    txt_product_name.send_keys(exc_product_name)

    txt_product_quantity = driver.find_element_by_name('product_quantity')
    txt_product_quantity.send_keys(exc_product_quantity)

    txt_product_price = driver.find_element_by_name('product_price')
    txt_product_price.send_keys(exc_product_price)

    txt_bntAdd = driver.find_element_by_name('bntAdd')
    txt_bntAdd.click()

login_to_web()

excelfile = load_workbook('sample.xlsx')
sheet = excelfile['DataBag']
maxrow = sheet.max_row
maxrow = maxrow+1

for i in range(2,maxrow):
    exc_product_name = sheet.cell(row=i,column=1).value
    exc_product_quantity = sheet.cell(row=i,column=2).value
    exc_product_price = sheet.cell(row=i,column=3).value
    add_product('https://noraphat.dev/POS_NAME/add-product.php',exc_product_name,exc_product_quantity,exc_product_price)